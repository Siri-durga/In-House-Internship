from django.shortcuts import render, redirect
from django.template.loader import render_to_string
from .forms import *
from django.utils import timezone
from django.contrib import messages
from .models import AddStudent, AddexamHall, Examallotment, AddFaculty, AdminAnnounce,AddTimeTable
from django.conf import settings
from django.core.mail import send_mail
from django.shortcuts import get_list_or_404
from itertools import chain
import secrets
import string
import random
from django.db.models import Q
import csv
from django.http import HttpResponse
from django.db import IntegrityError
import pandas as pd
from .forms import ExcelUploadForm
from django.contrib import messages
from .pdffile import *
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment
import re
from reportlab.lib.styles import ParagraphStyle
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from io import BytesIO
from .models import AddStudent

from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, PageBreak, KeepTogether
)
from reportlab.platypus import KeepTogether
# adminapp/views.py  (or wherever build_attendance_table lives)
import os                       # ← add this line
from io import BytesIO
from django.conf import settings
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Image, Spacer, PageBreak
)
# … the rest of your imports


from reportlab.lib.styles import getSampleStyleSheet

# To Delete Tables Data in Database
#AddTimeTable.objects.all().delete()
#Examallotment.objects.all().delete()
# AddexamHall.objects.all().delete()
#AddStudent.objects.all().delete()
#AddFaculty.objects.all().delete()
#Room.objects.all().delete()



def addtimetable(request):
    if request.method == 'POST':
        form = AddTimeTableForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('viewtimetable')  
    else:
        form = AddTimeTableForm()
    return render(request, 'addtimetable.html', {'form': form})

def viewtimetable(request):
    timetable_entries = AddTimeTable.objects.all()
    return render(request, 'viewtimetable.html', {'timetable_entries': timetable_entries})





# templates
INDEXPAGE = "index.html"
ADMINLOGINPAGE = "adminlogin.html"
ADMINHOMEPAGE = "adminhome.html"
ADDSTUDENTSPAGE = "addstudents.html"
ADDEXAMHALLSPAGE = "addexamhalls.html"
VIEWSTUDENTSPAGE = "viewstudents.html"
ADDFACULTYPAGE = "addfaculty.html"
VIEWFACULTYPAGE = "viewfaculty.html"
ADDANNOUNCEMENTPAGE = "addannouncement.html"
ADDTIMETABLEPAGE = "addtimetable.html"
VIEWTIMEPABLEPAGE = "viewtimetable.html"
# Create your views here.


def index(req):
    return render(req, INDEXPAGE)


def adminlogin(req):
    context = {}
    context['form'] = AdminlogForm()
    if req.method == "POST":
        form = AdminlogForm(req.POST)
        if form.is_valid():
            adminemail = form.cleaned_data['adminemail']
            adminpassword = form.cleaned_data['adminpassword']
            if adminemail == "admin@gmail.com" and adminpassword == "admin":
                req.session['adminemail'] = adminemail
                return render(req, ADMINHOMEPAGE)
            else:
                messages.warning(req, "Admin Credentials are not Valid......!")
                return render(req, ADMINLOGINPAGE, context)
    return render(req, ADMINLOGINPAGE, context)



def addstudents(req):
    context = {}
    if req.method == "POST":
        form = ExcelUploadForm(req.POST, req.FILES)
        if form.is_valid():
            excel_file = req.FILES['excel_file']
            if excel_file.name.endswith('.xlsx'):
                try:
                    df = pd.read_excel(excel_file)
                
                    for index, row in df.iterrows():
                        rollnumber = str(row['Roll.No'])
                        name = row['Student Name']
                        department = row['Department']
                        email = row['Email']
                        contact = re.sub(r'\D', '', str(row['Contact']))
                        year = row['Year']
                        semester = row['Semester']
                        #profile_url = row['Profile URL']

                        # Generate a random password
                        length = 8
                        characters = string.ascii_letters + string.digits
                        random_password = ''.join(secrets.choice(characters) for _ in range(length))

                        # Create a new instance of the AddStudent model
                        student = AddStudent.objects.create(
                            rollnumber=rollnumber,
                            name=name,
                            department=department,
                            email=email,
                            contact=contact,
                            year=year,
                            semester=semester,
                            #profile_url=profile_url,
                            password=random_password
                        )

                        # Sending email with login credentials
                        # subject = "Exam Details"
                        # cont = f'Dear {name},\n\n'
                        # m1 = f"Your Login Credentials:\nUsername: {rollnumber}\nPassword: {random_password}\n\n"
                        # m2 = "Thank you."
                        # m3 = "Regards,\nAdmin."

                        # email_from = settings.EMAIL_HOST_USER
                        # recipient_list = [email]
                        # message = cont + m1 + m2 + m3
                        # send_mail(subject, message, email_from, recipient_list, fail_silently=False)

                    messages.success(req, "Student details added successfully from Excel file.")
                    return redirect('addstudents')

                except Exception as e:
                    messages.error(req, f"Error processing Excel file: {e}")

            else:
                messages.error(req, "Please upload a valid Excel file.")

    else:
        form = ExcelUploadForm()

    context['form'] = form
    return render(req, 'addstudents.html', context)




def addexamhalls(request):
    if request.method == "POST":
        form = AddexamhallForm(request.POST)
        if form.is_valid():
            # ── pull cleaned data ───────────────────────────────
            date        = form.cleaned_data['Date']
            subject1    = form.cleaned_data['subject1']
            subject2    = form.cleaned_data['subject2']
            noofrooms   = form.cleaned_data['noofrooms']
            noofbenches = form.cleaned_data['noofbenches']
            students_per_bench = int(form.cleaned_data['students_per_bench'])

            selected_room_ids = request.POST.getlist('rooms')
            rooms = Room.objects.filter(id__in=selected_room_ids)
            selected_rooms = [room.room_number for room in rooms]

            # sanity-check: rooms chosen = number entered
            if len(selected_rooms) != noofrooms:
                form.add_error('rooms', f"Please select exactly {noofrooms} rooms.")
            else:
                total_benches = noofrooms * noofbenches
                total_seats_available = total_benches * students_per_bench
                total_students = AddStudent.objects.count()

                if total_seats_available < total_students:
                    messages.error(
                        request,
                        "Not enough seats available for the total number of students."
                    )
                else:
                    # ✅ DELETE previous exam hall records
                    AddexamHall.objects.all().delete()

                    # ✅ Create new exam hall
                    AddexamHall.objects.create(
                        date               = date,
                        subject1           = subject1,
                        subject2           = subject2,
                        noofrooms          = noofrooms,
                        noofbenches        = noofbenches,
                        total_benches      = total_benches,
                        total_seats        = total_seats_available,
                        rooms_list         = ",".join(selected_rooms),
                        students_per_bench = students_per_bench,
                    )

                    messages.success(request, "Exam hall added successfully.")
                    return redirect('addexamhalls')
        # else: form not valid → fall through to re-render with errors
    else:
        form = AddexamhallForm()

    rooms = Room.objects.all()
    return render(request, 'addexamhalls.html', {'form': form, 'rooms': rooms})


def delete(req, id):
    print(id)
    AddStudent.objects.filter(id=id).delete()
    return redirect("viewstudents")


def deletefaculty(req,id):
    AddFaculty.objects.filter(id=id).delete()
    return redirect("viewfaculty")



def setseatallotment(request):
    exam_hall_data = AddexamHall.objects.all()
    all_students = AddStudent.objects.order_by('id')

    # ✅ Clear previous seat allotment to avoid duplicates
    Examallotment.objects.all().delete()

    exam_halls = []
    for hall in exam_hall_data:
        exam_halls.append({
            'hall': hall,
            'total_benches': hall.total_benches,
            'total_rooms': hall.noofrooms,
            'benches_per_room': hall.noofbenches,
            'students_per_bench': hall.students_per_bench
        })

    allocated_students = set()

    for hall_info in exam_halls:
        hall = hall_info['hall']
        total_benches = hall_info['total_benches']
        total_rooms = hall_info['total_rooms']
        benches_per_room = hall_info['benches_per_room']
        students_per_bench = hall_info['students_per_bench']

        # Distribute students to benches
        for room_number in range(1, total_rooms + 1):
            allocated_seats = set()  # Track allocated seat numbers in the room
            for bench_number in range(1, benches_per_room + 1):
                branch_count = {}  # Track the count of students from each branch on the bench
                bench_students = []  # Track the students allocated to the current bench
                for seat_number in range(1, students_per_bench + 1):  # Seats per bench
                    student = None
                    # Find a student from a different branch
                    for candidate_student in all_students:
                        if candidate_student.id not in allocated_students and \
                                (branch_count.get(candidate_student.department, 0) < 1):
                            student = candidate_student
                            allocated_students.add(student.id)
                            branch_count[student.department] = branch_count.get(student.department, 0) + 1
                            break  # Allocate only one student per branch
                    if student:
                        # Ensure seat number is unique within the room
                        while seat_number in allocated_seats:
                            seat_number += 1  # Increment by 1
                        # Allocate the student to the seat
                        allocated_seats.add(seat_number)
                        bench_students.append(student)

                # If there are fewer students available than the specified number of students per bench
                if len(bench_students) < students_per_bench:
                    # Fill remaining seats with students from different branches
                    remaining_seats = students_per_bench - len(bench_students)
                    for _ in range(remaining_seats):
                        for candidate_student in all_students:
                            if candidate_student.id not in allocated_students:
                                # Add student from a different branch
                                if candidate_student.department not in branch_count:
                                    student = candidate_student
                                    allocated_students.add(student.id)
                                    branch_count[student.department] = branch_count.get(student.department, 0) + 1
                                    bench_students.append(student)
                                    break

                # Save allocation data to Examallotment model
                # Save allocation data to Examallotment model
                for student, seat_number in zip(bench_students, range(1, students_per_bench + 1)):
                    if student is None:
                        continue  # 🔒 Prevent crash

                    room_numbers = hall.rooms_list.split(',')
                    room_display = f"Room{room_numbers[room_number - 1]}"

                    # 🧠 Avoid duplicates
                    if not Examallotment.objects.filter(Student_Id=student.rollnumber, date=hall.date).exists():
                        Examallotment.objects.create(
                            Student_Id=student.rollnumber,
                            department=student.department,
                            RoomNo=room_display,
                            BenchNo=f"Bench{bench_number}",
                            SeatNumber=f"Seat{seat_number}",
                        
                            date=hall.date,
                    )

                    


                    # Prepare and send email to the student
                    # email_subject = "Exam Seat Allotment Details"
                    # email_message = f"Dear {student.name},\n\nYou have been allotted a seat for your upcoming exam.\n\nBranch: {student.department}\nRoom No: {room_display}\nBench No: {bench_number}\nSeat No: {seat_number}\nDate: {hall.date}\nStart Time: {hall.starttime}\nEnd Time: {hall.endtime}\n\nBest regards,\nAdmin"
                    # send_mail(email_subject, email_message, settings.EMAIL_HOST_USER, [student.email], fail_silently=False)


    return redirect("viewallotedstudents")



def viewallotedstudents(request):
    Exam_alloted_student = Examallotment.objects.all()
    return render(request, "viewallotedstudents.html", {'Exam_alloted_student': Exam_alloted_student})

def viewstudents(req):
    all_students = AddStudent.objects.all()
    return render(req, VIEWSTUDENTSPAGE, {'all_students': all_students})


def addfaculty(req):
    context = {}
    context['form'] = AddFacultyForm()
    if req.method == "POST":
        form = AddFacultyForm(req.POST)
        if form.is_valid():
            length = 8
            characters = string.ascii_letters + string.digits

            # Generate a random password
            random_password = ''.join(secrets.choice(characters) for _ in range(length))
            print("Random Password:", random_password)

            # Extracting form data
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            contact = form.cleaned_data['contact']
            branch = form.cleaned_data['branch']
            subject = form.cleaned_data['subject']
            semester = form.cleaned_data['semester']
            year = form.cleaned_data['year']
            #image = form.cleaned_data['image']
            #profilename = image.name

            try:
                # Attempt to save faculty member
                faculty = AddFaculty.objects.create(
                    name=name,
                    email=email,
                    contact=contact,
                    branch=branch,
                    subject=subject,
                    semester=semester,
                    year=year,
                    #image=image,
                    #profilename=profilename,
                    password=random_password
                )
            # # Mail Code
            #     # Sending email with login credentials
            #     subject = "Exam Details"
            #     cont = f'Dear {name}'
            #     KEY = f' Branch : {branch}\n'
            #     m1 = f"Your Login Credentials Username : {email}  & password {random_password}"
            #     m2 = "Thanking you"
            #     m3 = "Regards"
            #     m4 = "Admin."

            #     email_from = settings.EMAIL_HOST_USER
            #     recipient_list = [email]
            #     text = cont + '\n' + KEY + '\n' + m1 + '\n' + m2 + '\n' + m3 + '\n' + m4
            #     send_mail(subject, text, email_from, recipient_list, fail_silently=False)

                messages.success(req, "Faculty added successfully")
            except IntegrityError:
                # If email already exists, handle the exception
                messages.warning(req, "A faculty member with the same email already exists")

    return render(req, 'addfaculty.html', context)

#def addannouncement(req):
 #   two_days_content = timezone.now()-timezone.timedelta(days=2)
  #  messages_to_delete = AdminAnnounce.objects.filter(annuncementdate=two_days_content)
  #  all_messages = AdminAnnounce.objects.all()
  #  context = {}
   # context['form'] = AdminAnnouncement()
#
 #      form = AdminAnnouncement(req.POST)
  #      print(form.is_valid())
   #     if form.is_valid():
    #       adminemail = req.session['adminemail']
     #       data = AdminAnnounce(
     #           announcement=announcement,
     #           senderemail=adminemail
     #       )
     #       data.save()

     #       # Correct syntax for passing context to the template
     #       return render(req, ADDANNOUNCEMENTPAGE, {'form': AdminAnnouncement(), 'all_messages': all_messages})

   # return render(req, ADDANNOUNCEMENTPAGE, {'form': AdminAnnouncement(), 'all_messages': all_messages})



def viewfaculty(req):
    all_faculty = AddFaculty.objects.all()
    return render(req, VIEWFACULTYPAGE, {'all_faculties': all_faculty})



def download_details(req):
    # Replace YourModel with your actual model
    details_data = Examallotment.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="details.xlsx"'

    # Create a new Excel workbook and add a worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Add college name, department, and venue information
    worksheet.merge_cells('A1:I1')  # Merge cells for college name
    college_name_cell = worksheet['A1']
    college_name_cell.value = "Aditya Unversity"
    college_name_cell.font = Font(bold=True, size=18)
    college_name_cell.alignment = Alignment(horizontal='center')


    worksheet.merge_cells('A2:I2')  # Merge cells for venue
    venue_cell = worksheet['A2']
    venue_cell.value = "Venue: BGB"
    venue_cell.font = Font(bold=True)
    venue_cell.alignment = Alignment(horizontal='center')

    # Add header row for exam allotment
    header_row = ['Branch','RoomNo', 'BenchNo', 'SeatNumber',
                   'Student_Id', 'Date', 'Start Time', 'End Time']
    worksheet.append(header_row)

    # Apply bold font to the header row
    for cell in worksheet.iter_rows(min_row=worksheet.max_row, max_row=worksheet.max_row, min_col=1, max_col=len(header_row)):
        for c in cell:
            c.font = Font(bold=True)

    # Write data rows for exam allotment
    for detail in details_data:
        data_row = [detail.department,detail.RoomNo, detail.BenchNo, detail.SeatNumber, 
                    detail.Student_Id, detail.date, detail.starttime, detail.endtime]
        worksheet.append(data_row)

    # Save the workbook to the response
    workbook.save(response)

    return response

def attendance_sheet_home(request):
    """
    Generate a single PDF containing attendance sheets for every room.
    Each room gets up to two sections: JUNIORS (23-series) and SENIORS (20-series).
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=36,
        rightMargin=36,
        topMargin=30,
        bottomMargin=20
    )

    elements = []
    styles = getSampleStyleSheet()

    # All distinct room labels in Examallotment, e.g.  "Room103"
    all_rooms = (
        Examallotment.objects
        .values_list('RoomNo', flat=True)
        .distinct()
        .order_by('RoomNo')
    )

    first_section = True   # to skip the very first PageBreak

    for room_label in all_rooms:
        # -------- find the AddexamHall row for THIS room ----------
        #   We pull the numeric portion (e.g. 103) and search rooms_list.
        numeric_room = re.sub(r'[^0-9]', '', room_label)
        hall = (
            AddexamHall.objects
            .filter(rooms_list__icontains=numeric_room)
            .first()
        )

        # Fetch every student who ended up in that room
        student_ids = (
            Examallotment.objects
            .filter(RoomNo=room_label)
            .values_list('Student_Id', flat=True)
            .distinct()
        )
        students = (
            AddStudent.objects
            .filter(rollnumber__in=student_ids)
            .order_by('rollnumber')
        )

        juniors = [s for s in students if s.rollnumber.startswith('23')]
        seniors = [s for s in students if s.rollnumber.startswith('20')]

        # -------- build JUNIOR section ----------
        if juniors:
            elements += build_attendance_table(
                hall,
                juniors,
                room_label,
                "JUNIORS",
                styles,
                add_page_break=not first_section
            )
            first_section = False     # first section handled

        # -------- build SENIOR section ----------
        if seniors:
            elements += build_attendance_table(
                hall,
                seniors,
                room_label,
                "SENIORS",
                styles,
                add_page_break=not first_section
            )
            first_section = False

    # Finalise PDF
    doc.build(elements)
    buffer.seek(0)
    return HttpResponse(buffer, content_type="application/pdf")

from reportlab.platypus import Paragraph, Table, TableStyle, Spacer, PageBreak
from reportlab.lib import colors

def build_attendance_table(
        hall,               # ← NEW: the AddexamHall instance
        student_list,
        room,
        label,
        styles,
        add_page_break=True
):
    """
    Build a single attendance-sheet section for the given room / label.
    """

    elements = []

    # ---- page break between sections ----
    if add_page_break:
        elements.append(PageBreak())
        
    logo_path = os.path.join(
        settings.BASE_DIR,              # project root
        'static',                       #  ← adjust if your static dir differs
        'adlogo.png'                     #  ← replace with your file name
        )

    # 2) build an Image flowable (tweak size as you like)
    logo = Image(logo_path, width=350, height=80)  # pick size you like
    logo.hAlign = "CENTER"  # px ≈ points here

    # 3) wrap it in a 1-cell table and center it
    #logo_tbl = Table([[logo]], colWidths=[500])    # same width you use later
    #logo_tbl.hAlign = 'CENTER'

    # 4) add to elements *before* the title
    elements.append(logo)
    elements.append(Spacer(1, 6)) 

    # ---- Title & department ----
    #elements.append(Paragraph("<b>ADITYA UNIVERSITY</b>", styles['Title']))
    #elements.append(Spacer(1, 12))

    dept = (
        "Computer Science and Engineering"
        if label == "JUNIORS"
        else "Electronics and Communication Engineering"
    )
    centered_heading = ParagraphStyle(
    name='CenteredHeading',
    parent=styles['Heading2'],
    alignment=1  # 0=left, 1=center, 2=right, 4=justify
    )

    elements.append(Paragraph(f"<b>Department of {dept}</b>", centered_heading))
    elements.append(Spacer(1, 6))

    # ---- Left / right metadata ----
    left_metadata = [
        [Paragraph("<b>Venue:</b> BGB", styles['BodyText'])],
        [Paragraph(f"<b>Room Number:</b> {room}", styles['BodyText'])],
    ]
    left_table = Table(left_metadata, hAlign='LEFT')

    semester = "IV Sem" if label == "JUNIORS" else "VIII Sem"
    right_metadata = [
        [
            Paragraph(
                f"<b>Date:</b> {hall.date.strftime('%d-%m-%Y')}",
                styles['BodyText']
            )
        ],
        [Paragraph(f"<b>Semester:</b> {semester}", styles['BodyText'])],
    ]
    right_table = Table(right_metadata, hAlign='RIGHT')

    combined = Table([[left_table, right_table]], colWidths=[350, 150])
    combined.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    elements.extend([combined, Spacer(1, 12)])

    # ---- Attendance Sheet heading ----
    elements.append(Paragraph("<b>Attendance Sheet</b>", styles['Heading2']))
    elements.append(Spacer(1, 12))

    # ---- Main table ----
    table_data = [["S.No", "Roll No", "Name", "Signature"]]
    for idx, stu in enumerate(student_list, start=1):
        table_data.append([idx, stu.rollnumber, stu.name, ""])

    att_table = Table(
        table_data,
        colWidths=[50, 120, 250, 80],    # wider Name column
        repeatRows=1
    )
    att_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID',       (0, 0), (-1, -1), 1, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(att_table)

    # ---- Invigilator signature lines ----
    elements.append(Spacer(1, 36))
    signature_tbl = Table([
        ["", "", "__________________________", "__________________________"],
        ["", "", "Invigilator 1 Signature",   "Invigilator 2 Signature"],
    ], colWidths=[50, 150, 150, 150])
    signature_tbl.setStyle(TableStyle([
        ('ALIGN', (2, 0), (3, 0), 'CENTER'),
        ('ALIGN', (2, 1), (3, 1), 'CENTER'),
        ('FONTSIZE', (2, 1), (3, 1), 10),
        ('TOPPADDING', (2, 0), (3, 0), 20),
    ]))
    elements.append(signature_tbl)

    return elements